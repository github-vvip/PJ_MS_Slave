"""
Excel 配置表解析模块
从"订单软硬件配置表"Excel 文件中解析 11 个扩展字段
依赖：openpyxl (>=3.1.5) — 支持 CellRichText 富文本解析
"""
import re
import openpyxl
from openpyxl.cell.rich_text import CellRichText

# ============================================================
# 屏幕型号兜底映射表
# ============================================================
SCREEN_MODEL_MAP = {
    '10.1': 'CL50：JYM1015853281BB',
    '15.6': 'CL173：QSJ156CS02-1',
}


# ============================================================
# 工具函数
# ============================================================

def _cell_text(ws, cell_ref):
    """安全读取单元格纯文本内容"""
    cell = ws[cell_ref]
    if cell.value is None:
        return ''
    if isinstance(cell.value, CellRichText):
        return ''.join(block.text for block in cell.value)
    return str(cell.value)


def _is_isolated_keyword(block_text, keyword):
    """
    检查 keyword 是否在 block_text 中以"独立词"形式出现
    前后必须是 □/■/空格/制表符/字符串边界等分隔符，避免子串误匹配
    （例如：避免 'COB' 中的 'CO' 误匹配，或 'COF' 在 '□COB  □COF' 中误匹配 'COB'）
    """
    if not block_text or not keyword:
        return False
    pattern = r'(?:^|[□\s])' + re.escape(keyword) + r'(?:[□\s]|$)'
    return re.search(pattern, block_text) is not None


def _has_red_font(cell, text):
    """
    检测单元格中是否存在红色字体的指定文本
    返回 True 表示该文本以红色字体呈现（视为选中状态）
    优先级: 富文本段落 > 整单元格字体颜色
    要求: text 必须是独立词（前后是 □/■/空格/边界），避免子串误匹配
    """
    # 情况 1: 富文本 CellRichText → 只检查具体 text block 的字体颜色
    if isinstance(cell.value, CellRichText):
        for block in cell.value:
            if not _is_isolated_keyword(block.text, text):
                continue
            font = block.font
            if font and font.color:
                try:
                    rgb = str(font.color.rgb) if font.color.rgb else ''
                except Exception:
                    continue
                rgb_upper = rgb.upper().replace('0X', '')
                if len(rgb_upper) == 8 and rgb_upper.startswith('00'):
                    rgb_upper = rgb_upper[2:]
                if 'FF0000' in rgb_upper:
                    return True
        return False

    # 情况 2: 纯文本单元格 → 检查整单元格字体颜色
    plain = str(cell.value or '')
    if _is_isolated_keyword(plain, text) and cell.font and cell.font.color:
        try:
            rgb = str(cell.font.color.rgb) if cell.font.color.rgb else ''
        except Exception:
            rgb = ''
        rgb_upper = rgb.upper().replace('0X', '')
        if len(rgb_upper) == 8 and rgb_upper.startswith('00'):
            rgb_upper = rgb_upper[2:]
        if 'FF0000' in rgb_upper:
            return True
    return False


def _has_text_in_region(ws, area_start, area_end, keyword):
    """
    在指定区域内搜索包含关键词的单元格
    返回: 匹配的单元格对象，或 None
    area_start: 如 'G9'
    area_end: 如 'G14'
    """
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(f'{area_start}:{area_end}')
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if keyword in str(cell.value or ''):
                return cell
    return None


def _right_cell(ws, cell):
    """取同行右侧相邻单元格"""
    return ws.cell(row=cell.row, column=cell.column + 1)


def _square_after_text(text):
    """提取第一个 ■ 之后的文本（直到空格或末尾）"""
    if not text:
        return ''
    idx = text.find('■')
    if idx < 0:
        return ''
    rest = text[idx + 1:].strip()
    # 取到第一个空格或结束
    parts = rest.split(' ', 1)
    return parts[0].strip()


def _extract_after(text, keyword):
    """提取关键词之后的连续字母数字内容"""
    idx = text.find(keyword)
    if idx < 0:
        return ''
    rest = text[idx + len(keyword):].strip()
    val = ''
    for ch in rest:
        if ch.isalnum() or ch in '.*/-':
            val += ch
        elif val:
            break
    return val


# ============================================================
# 字段解析函数
# ============================================================

def parse_model(ws):
    """
    型号 — 检索区域 D18, D19, D20, D21
    遍历单元格，找包含"型号"的 → 读取同行右侧相邻单元格
    """
    for row in [18, 19, 20, 21]:
        cell_text = _cell_text(ws, f'D{row}')
        if '型号' in cell_text:
            return _cell_text(ws, f'E{row}')
    return ''


def parse_brand(ws):
    """
    厂商（覆盖式） — 检索区域 D18, D19, D20
    遍历三个单元格，找包含"厂商"的 → 读取同行右侧相邻单元格
    若解析到值则覆盖基础字段中的厂商；否则保持原值不变
    """
    for row in [18, 19, 20]:
        cell_text = _cell_text(ws, f'D{row}')
        if '厂商' in cell_text:
            val = _cell_text(ws, f'E{row}')
            if val:
                return val
    return None  # None 表示不覆盖


def _parse_led_pir(ws, keyword):
    """
    LED/PIR 共用解析逻辑 — 检索区域 G9:G14
    步骤 1: 找包含 keyword 的单元格 → 取右侧相邻 H 列
    步骤 2: 检查富文本红色字体的"有"/"无"
    步骤 3: 回退到 □/■ 方块字符索引判断
    """
    for row in range(9, 15):
        cell_text = _cell_text(ws, f'G{row}')
        if keyword not in cell_text:
            continue
        target = ws[f'H{row}']
        # 步骤 2: 红色字体检测
        if _has_red_font(target, '有'):
            return '有'
        if _has_red_font(target, '无'):
            return '无'
        # 步骤 3: 方块字符 — ■ 后面的文本为选中值
        plain = str(target.value or '')
        idx_filled = plain.find('■')
        if idx_filled >= 0:
            after = plain[idx_filled + 1:].strip()
            if after.startswith('有'):
                return '有'
            elif after.startswith('无'):
                return '无'
        # 无 ■ 或无法识别 → 默认无
        return '无'
    return '无'


def parse_led(ws):
    """LED 字段解析"""
    return _parse_led_pir(ws, 'RGB')


def parse_pir(ws):
    """PIR 字段解析"""
    return _parse_led_pir(ws, 'PIR')


def _detect_light_sensor_model(ws, project_name):
    """
    光感型号判断子流程
    优先级 1: 项目名称包含 ADC → ADCF3, 包含 3311 → STK3311
    优先级 2: 搜索 B3:J26 区域
    """
    pn = (project_name or '').upper()
    if 'ADC' in pn:
        return 'ADCF3'
    if '3311' in pn:
        return 'STK3311'
    # 搜索 B3:J26
    found = []
    for row in range(3, 27):
        for col in range(2, 11):  # B=2, J=10
            val = str(ws.cell(row=row, column=col).value or '').upper()
            if 'ADC' in val:
                found.append('ADCF3')
            if '3311' in val:
                found.append('STK3311')
    if found:
        return found[-1]  # 取最后出现的
    return '有'  # 仅标记有，不指定型号


def parse_light_sensor(ws, project_name=''):
    """
    光感 — 检索区域 G9:G14，最复杂
    步骤 1: 找包含"光感"的单元格 → 取右侧 H 列
    步骤 2: 红色字体检测
    步骤 3: □/■ 索引判断
    步骤 4: 光感型号子流程
    """
    for row in range(9, 15):
        cell_text = _cell_text(ws, f'G{row}')
        if '光感' not in cell_text:
            continue
        target = ws[f'H{row}']
        # 红色字体
        if _has_red_font(target, '无'):
            return '无'
        if _has_red_font(target, '有'):
            return _detect_light_sensor_model(ws, project_name)
        # 方块字符 — ■ 后面的文本为选中值
        plain = str(target.value or '')
        idx_filled = plain.find('■')
        if idx_filled >= 0:
            after = plain[idx_filled + 1:].strip()
            if after.startswith('有'):
                return _detect_light_sensor_model(ws, project_name)
            elif after.startswith('无'):
                return '无'
        return '无'
    return '无'


def parse_wifi(ws):
    """
    WiFi — 检索区域 A8:A14
    步骤 1: 找包含"网络"或"WiFi"的单元格 → 取右侧 B 列
    步骤 2: 红色字体检测 "5G" / "2.4G"
    步骤 3: ■ 后文本判断
    """
    for row in range(8, 15):
        cell_text = _cell_text(ws, f'A{row}')
        if '网络' not in cell_text and 'WiFi' not in cell_text:
            continue
        target = ws.cell(row=row, column=2)  # B 列
        # 红色字体检测
        if _has_red_font(target, '5G'):
            return '5G'
        if _has_red_font(target, '2.4G'):
            return '2.4G'
        # ■ 后文本判断
        plain = str(target.value or '')
        after = _square_after_text(plain)
        if '5G' in after:
            return '5G'
        if '单' in after:
            return '2.4G'
        if '双' in after:
            return '5G'
        return '2.4G'
    return ''


def parse_screen_size(ws):
    """
    屏幕尺寸 — 检索区域 B8
    提取"尺寸"后的数字 → 提取"分辨率"后的数字 → 拼接
    """
    text = _cell_text(ws, 'B8')
    size = ''
    resolution = ''

    # 提取尺寸值
    idx = text.find('尺寸')
    if idx >= 0:
        rest = text[idx + 2:]
        digit_start = -1
        for i, ch in enumerate(rest):
            if ch.isdigit() or ch == '.':
                if digit_start < 0:
                    digit_start = i
            elif digit_start >= 0:
                size = rest[digit_start:i]
                break
        if digit_start >= 0 and not size:
            size = rest[digit_start:]

    # 提取分辨率
    idx = text.find('分辨率')
    if idx >= 0:
        rest = text[idx + 3:]
        digit_start = -1
        for i, ch in enumerate(rest):
            if ch.isdigit() or ch == '*':
                if digit_start < 0:
                    digit_start = i
            elif digit_start >= 0:
                resolution = rest[digit_start:i]
                break
        if digit_start >= 0 and not resolution:
            resolution = rest[digit_start:]

    if size:
        return f'{size}寸{resolution}' if resolution else f'{size}寸'
    return ''


def parse_screen_model(ws, screen_size=''):
    """
    屏幕型号 — 检索区域 B8
    步骤 1: 找到"型号"后第一个字母或数字，读取直到遇到空格/中文字符或字符串结束
    步骤 2: 兜底映射
    """
    text = _cell_text(ws, 'B8')
    idx = text.find('型号')
    if idx >= 0:
        rest = text[idx + 2:]
        # 跳过冒号和空格，找到第一个字母或数字
        start = None
        for i, ch in enumerate(rest):
            if ch.isalnum():
                start = i
                break
        if start is not None:
            val = ''
            for ch in rest[start:]:
                # 遇到空格或中文字符则停止
                if ch == ' ' or ('一' <= ch <= '鿿'):
                    break
                val += ch
            # 去掉尾部非字母数字字符（如 "-"、"："）
            while val and not val[-1].isalnum():
                val = val[:-1]
            if val:
                return val

    # 兜底映射
    if screen_size:
        for key, model in SCREEN_MODEL_MAP.items():
            if screen_size.startswith(key):
                return model
    return '中性'


def parse_tp(ws):
    """
    TP — 检索区域 B9
    步骤 1: 提取第一个 ■ 后的 3 个字母（接口类型）
    步骤 1b: 如果没有 ■，按 □ 切分，找独立出现且为红色字体的关键字
    步骤 2: 提取"型号"后的内容
    步骤 3: 拼接
    """
    cell = ws['B9']
    text = str(cell.value or '')
    interface = ''
    model = ''

    # 接口类型：■ 优先
    filled_idx = text.find('■')
    if filled_idx >= 0:
        after = text[filled_idx + 1:].strip()
        letters = ''
        for ch in after:
            if ch.isalpha():
                letters += ch
            else:
                break
        if letters:
            interface = letters[:3]
    else:
        # 无 ■，按 □ 切分各选项，取独立出现且为红色字体的关键字
        TP_KEYWORDS = ['COF', 'COB', 'USB']
        segments = text.split('□')
        for seg in segments:
            seg = seg.strip()
            if not seg:
                continue
            for kw in TP_KEYWORDS:
                if seg.startswith(kw) and _has_red_font(cell, kw):
                    interface = kw
                    break
            if interface:
                break

    # 型号
    idx = text.find('型号')
    if idx >= 0:
        rest = text[idx + 2:].strip()
        val = ''
        for ch in rest:
            if ch.isalnum() or ch in '-_':
                val += ch
            elif val:
                break
        model = val

    if interface and model:
        return f'{interface}-{model}'
    if interface:
        return interface
    if model:
        return model
    return ''


def _is_block_red(block, cell=None):
    """
    判断富文本块（或整个单元格）的字体颜色是否为红色
    返回 True 表示该块字体包含红色
    """
    font = getattr(block, 'font', None) or (cell.font if cell is not None else None)
    if not font or not font.color:
        return False
    try:
        rgb = str(font.color.rgb) if font.color.rgb else ''
    except Exception:
        return False
    rgb_upper = rgb.upper().replace('0X', '')
    if len(rgb_upper) == 8 and rgb_upper.startswith('00'):
        rgb_upper = rgb_upper[2:]
    return 'FF0000' in rgb_upper


def _extract_chinese(text):
    """提取文本中所有连续汉字段，按非汉字切分"""
    if not text:
        return []
    result = []
    current = ''
    for ch in text:
        if '一' <= ch <= '鿿':
            current += ch
        else:
            if current:
                result.append(current)
                current = ''
    if current:
        result.append(current)
    return result


def _get_red_texts_from_xml(cell, file_path):
    """
    当 openpyxl 无法解析为 CellRichText（返回纯 str）时，
    直接解析 xlsx 底层 XML，提取红色字体的文本段落。
    原因：openpyxl 在 data_only=True 时丢失 _archive 引用，
    且共享字符串首段无 rPr 时会丢弃全部富文本信息，
    但 XML 中实际存在 <color rgb="FFFF0000"/> 标记。
    返回: 红色文本列表，如 ['客户私模']
    """
    import xml.etree.ElementTree as ET
    import zipfile
    import re

    try:
        cell_ref = cell.coordinate  # e.g. 'B25'
        col_letter = re.match(r'([A-Z]+)', cell_ref).group(1)
        row_num = int(re.match(r'[A-Z]+(\d+)', cell_ref).group(1))

        with zipfile.ZipFile(file_path, 'r') as z:
            # 步骤1: 读取 sheet1.xml，获取单元格的共享字符串索引
            if 'xl/worksheets/sheet1.xml' not in z.namelist():
                return []
            sheet_data = z.read('xl/worksheets/sheet1.xml')
            sheet_root = ET.fromstring(sheet_data)
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            ss_index = None
            for row_elem in sheet_root.findall('.//ns:row', ns):
                r = int(row_elem.get('r', '0'))
                if r != row_num:
                    continue
                for c_elem in row_elem.findall('ns:c', ns):
                    if c_elem.get('r', '') == cell_ref:
                        cell_type = c_elem.get('t', '')
                        if cell_type == 's':  # shared string
                            v_elem = c_elem.find('ns:v', ns)
                            if v_elem is not None:
                                ss_index = int(v_elem.text)
                        break
                break

            if ss_index is None:
                return []

            # 步骤2: 读取 sharedStrings.xml，获取该索引的红色文本
            if 'xl/sharedStrings.xml' not in z.namelist():
                return []
            ss_data = z.read('xl/sharedStrings.xml')
            ss_root = ET.fromstring(ss_data)

            si_list = ss_root.findall('ns:si', ns)
            if ss_index >= len(si_list):
                return []
            si = si_list[ss_index]

            red_texts = []
            for r_elem in si.findall('ns:r', ns):
                rpr = r_elem.find('ns:rPr', ns)
                is_red = False
                if rpr is not None:
                    color_elem = rpr.find('ns:color', ns)
                    if color_elem is not None:
                        rgb = color_elem.get('rgb', '')
                        if 'FF0000' in rgb.upper():
                            is_red = True
                if is_red:
                    t_elem = r_elem.find('ns:t', ns)
                    if t_elem is not None and t_elem.text:
                        text = t_elem.text.strip()
                        if text:
                            red_texts.append(text)
            return red_texts
    except Exception:
        return []


def parse_shell(ws, file_path=None):
    """
    壳 — 检索区域 A22, A23, A24, A25
    规则 1: 找包含"模"或"具"或"壳"的 → 取右侧 B 列 → ■ 后文本（跳过"其它"/"其他"）
    规则 2 (方案 A): ■ 未提取到时，遍历 B 列所有富文本块，
            找到字体为红色、且包含独立汉字段的块 → 提取该块的汉字段
            不依赖 □ 切分，不依赖位置先后，可处理"客户私模"等空格后的红色文字
    file_path: 当 openpyxl 丢失富文本信息时，用于直接解析 xlsx XML 底层
    """
    SKIP_PREFIXES = ['其它', '其他']

    for row in [21, 22, 23, 24, 25]:
        cell_text = _cell_text(ws, f'A{row}')
        if '模' not in cell_text and '具' not in cell_text and '壳' not in cell_text:
            continue

        cell = ws[f'B{row}']
        target_text = _cell_text(ws, f'B{row}')

        # 规则 1: 找到第一个 ■
        filled_idx = target_text.find('■')
        if filled_idx >= 0:
            after = target_text[filled_idx + 1:].strip()
            for prefix in SKIP_PREFIXES:
                if after.startswith(prefix):
                    after = after[len(prefix):].strip()
            parts = after.split(' ', 1)
            val = parts[0].strip().replace('：', '').replace(':', '').replace('　', '')
            if val:
                return val

        # 规则 2 (方案 A): 遍历所有富文本块，找红色字体的汉字段
        if isinstance(cell.value, CellRichText):
            for block in cell.value:
                if not _is_block_red(block):
                    continue
                block_text = block.text
                # 提取块中所有汉字段
                chinese_segments = _extract_chinese(block_text)
                if not chinese_segments:
                    continue
                # 取第一个非黑名单的汉字段
                for seg in chinese_segments:
                    val = seg.replace('：', '').replace(':', '').replace('　', '').strip()
                    skip = False
                    for prefix in SKIP_PREFIXES:
                        if val.startswith(prefix):
                            skip = True
                            break
                    if not skip and val:
                        return val

        # 规则 2b: openpyxl 未解析为 CellRichText 时，直接解析 xlsx XML 提取红色文本
        if not isinstance(cell.value, CellRichText) and isinstance(cell.value, str) and file_path:
            red_texts = _get_red_texts_from_xml(cell, file_path)
            for red_text in red_texts:
                chinese_segments = _extract_chinese(red_text)
                for seg in chinese_segments:
                    val = seg.replace('：', '').replace(':', '').replace('　', '').strip()
                    skip = False
                    for prefix in SKIP_PREFIXES:
                        if val.startswith(prefix):
                            skip = True
                            break
                    if not skip and val:
                        return val

        # 兜底：整单元格字体为红色 → 取所有汉字段中第一个非黑名单
        if cell.value and cell.font and _is_block_red(None, cell):
            chinese_segments = _extract_chinese(target_text)
            for seg in chinese_segments:
                val = seg.replace('：', '').replace(':', '').replace('　', '').strip()
                skip = False
                for prefix in SKIP_PREFIXES:
                    if val.startswith(prefix):
                        skip = True
                        break
                if not skip and val:
                    return val
    return ''


def parse_date(ws):
    """
    立项时间 — 检索区域 H2
    情况 1: Excel 日期类型 → 格式化为 yyyy/M/d
    情况 2: 文本格式 → 解析年月日
    """
    cell = ws['H2']
    val = cell.value
    if val is None:
        return None

    # 情况 1: Excel datetime/date 类型
    if hasattr(val, 'strftime'):
        return val.strftime('%Y/%-m/%-d') if hasattr(val, 'strftime') else None

    # 情况 2: 文本格式
    text = str(val)
    year = ''
    month = '1'
    day = '1'

    idx = text.find('年')
    if idx >= 0:
        year = text[:idx].strip()
    idx = text.find('月')
    if idx >= 0:
        # 从"年"后到"月"前
        year_end = text.find('年')
        if year_end >= 0:
            month = text[year_end + 1:idx].strip()
    idx = text.find('日')
    if idx >= 0:
        month_end = text.find('月')
        if month_end >= 0:
            day = text[month_end + 1:idx].strip()
        else:
            year_end = text.find('年')
            if year_end >= 0:
                day = text[year_end + 1:idx].strip()

    # 去掉前导零
    month = str(int(month)) if month.isdigit() else '1'
    day = str(int(day)) if day.isdigit() else '1'

    if year and year.isdigit():
        return f'{year}/{month}/{day}'
    return None


def parse_remarks(ws):
    """
    备注 — 多段拼接，以"；"分隔
    段 1: H5 ■ 后的文本
    段 2: G8:G14 中"重力感应"右侧 ■ 后的文本
    段 3: G8:G14 中"GMS"右侧 ■ 后的文本
    """
    parts = []

    # 段 1: H5
    h5_text = _cell_text(ws, 'H5')
    val = _square_after_text(h5_text)
    if val:
        parts.append(val)

    # 段 2: G8:G14 中"重力感应"右侧 ■ 后的文本
    for row in range(8, 15):
        cell_text = _cell_text(ws, f'G{row}')
        if '重力感应' in cell_text:
            target_text = _cell_text(ws, f'H{row}')
            val = _square_after_text(target_text)
            if val:
                parts.append(f'重力感应：{val}')

    # 段 3: G8:G16 中"GMS"右侧 ■ 后的文本
    for row in range(8, 17):
        cell_text = _cell_text(ws, f'G{row}')
        if 'GMS' in cell_text.upper():
            target_text = _cell_text(ws, f'H{row}')
            val = _square_after_text(target_text)
            if val:
                parts.append(f'GMS：{val}')

    return '；'.join(parts) if parts else ''


def parse_launcher_from_excel(ws):
    """
    Launcher — 检索区域 B21:B24
    仅在基础 Launcher 为 OM 时调用
    步骤 1: 检查红色字体
    步骤 2: ■ 后文本 → 关键词映射
    """
    LAUNCHER_RED_RULES = [
        ('FrameO', 'FM'),
        ('Photo', 'WP'),
        ('Uhale', 'UH'),
        ('MTKCalen', 'CT'),
        ('kairos', 'CT'),
    ]

    # ■ 后文本 → Launcher 映射表（不区分大小写）
    LAUNCHER_MAP = {
        'photo': 'WP',
        'frameo': 'FM',
        'whaleframely': 'WF',
        'uhale': 'UH',
        'timer': 'CT',
        'manufacturer': 'CT',
        'mtkcal': 'CT',
        'kairos': 'CT',
        '智象日历': 'WF',
    }

    for row in range(21, 25):
        cell = ws[f'B{row}']
        # 步骤 1: 红色字体检测
        for keyword, launcher in LAUNCHER_RED_RULES:
            if _has_red_font(cell, keyword):
                return launcher

        # 步骤 2: ■ 后文本判断 → 关键词映射
        text = str(cell.value or '')
        filled_idx = text.find('■')
        if filled_idx >= 0:
            after = text[filled_idx + 1:].strip()
            # 跳过"其它"二字
            if after.startswith('其它'):
                after = after[2:].strip()
            # 读取直到遇到空格或结束
            parts = after.split(' ', 1)
            val = parts[0].strip().lower()
            if val:
                # 关键词映射
                for keyword, launcher in LAUNCHER_MAP.items():
                    if keyword in val or keyword == val:
                        return launcher
                # 未匹配到任何关键词，不覆盖（沿用雷达检索值 OM）
                return None

    return None  # 不覆盖


# ============================================================
# 主入口
# ============================================================

def parse_excel_config(file_path, project_name=''):
    """
    解析 Excel 配置表，返回所有扩展字段字典
    参数:
        file_path: Excel 文件路径
        project_name: 项目名称（用于光感型号子流程）
    返回:
        dict: {model, brand, pir, led, light_sensor, wifi,
               screen_size, screen_model, tp, shell,
               project_establish_date, remarks}
        注意: brand 为 None 表示不覆盖基础字段
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f'无法打开 Excel 文件: {e}')

    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError('Excel 文件没有活动工作表')

    try:
        result = {
            'model': parse_model(ws),
            'brand': parse_brand(ws),
            'pir': parse_pir(ws),
            'led': parse_led(ws),
            'light_sensor': parse_light_sensor(ws, project_name),
            'wifi': parse_wifi(ws),
            'screen_size': parse_screen_size(ws),
            'screen_model': '',
            'tp': parse_tp(ws),
            'shell': parse_shell(ws, file_path),
            'project_establish_date': parse_date(ws),
            'remarks': parse_remarks(ws),
            'launcher': parse_launcher_from_excel(ws),
        }
        # screen_model 依赖 screen_size
        result['screen_model'] = parse_screen_model(ws, result['screen_size'])
    finally:
        wb.close()

    return result
